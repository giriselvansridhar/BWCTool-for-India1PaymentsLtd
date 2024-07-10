import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk, scrolledtext
import json
import os
import pandas as pd
import csv
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from FileValidation import FileNameValidator
from FileParsing import FilePraser
from Models.JsonPropertyReader import JsonPropertyReader

class ConsolidationApp:
    def __init__(self, root):
        self.root = root
        self.root.title("BSC Tool")
        self.root.geometry("500x400")
        self.root.configure(bg='#FFFFFF')  # White background
        
        # Add a banner for the company name
        self.banner_frame = tk.Frame(self.root, bg='#E3F2FD')
        self.banner_frame.pack(fill=tk.X)

        self.company_label = ttk.Label(self.banner_frame, text="India1 Payments Ltd", font=("Arial", 18, "bold"), background='#E3F2FD', foreground='#0D47A1')
        self.company_label.pack(pady=10)
        
        self.progress_var = tk.DoubleVar()
        self.status_var = tk.StringVar()

        self.style = ttk.Style()
        self.style.theme_use("clam")

        self.style.configure("TButton", padding=6, relief="flat", background="#4CAF50", foreground="white", font=("Arial", 10))
        self.style.map("TButton", background=[('active', '#45a049')])
        self.style.configure("Rounded.TButton", borderwidth=0, relief="flat", background="#4CAF50", foreground="white", font=("Arial", 10), padding=6)
        self.style.map("Rounded.TButton", background=[('active', '#45a049')])

        self.style.configure("TProgressbar", thickness=10, background="#2196F3", troughcolor="#BBDEFB")

        self.create_widgets()

    def create_widgets(self):
        self.heading_label = ttk.Label(self.root, text="BSC Tool", font=("Arial", 16, "bold"), background='#FFFFFF', foreground='#0D47A1')
        self.heading_label.pack(pady=10)

        # Enhance the "Consolidate Excel" button
        self.load_button = ttk.Button(self.root, text="Consolidate Excel", command=self.start_consolidation, style="Rounded.TButton")
        self.load_button.pack(pady=10, fill=tk.X, padx=20)

        self.progress_bar = ttk.Progressbar(self.root, variable=self.progress_var, maximum=100, mode='determinate', length=480)
        self.progress_bar.pack(pady=10, padx=10)

        self.status_label = ttk.Label(self.root, textvariable=self.status_var, anchor='center', background='#FFFFFF', foreground='#0D47A1')
        self.status_label.pack(pady=10, fill=tk.X, padx=20)

        self.bottom_frame = tk.Frame(self.root, bg='#FFFFFF')
        self.bottom_frame.pack(side=tk.BOTTOM, fill=tk.X)

        self.close_button = ttk.Button(self.bottom_frame, text="Close", command=self.root.quit, style="Rounded.TButton")
        self.close_button.pack(side=tk.RIGHT, padx=10, pady=10)

        self.instruction_button = ttk.Button(self.bottom_frame, text="Instructions", command=self.show_instructions, style="Rounded.TButton")
        self.instruction_button.pack(side=tk.RIGHT, padx=10, pady=10)


        # Update the copyright label
        self.copy_right_label = ttk.Label(self.root, text="© India1 Payments Ltd", anchor='e', background='#FFFFFF', foreground='#0D47A1')
        self.copy_right_label.pack(side=tk.BOTTOM, pady=5, fill=tk.X)

    def start_consolidation(self):
        self.progress_var.set(0)
        self.status_var.set("Starting consolidation...")

        ready_directory = 'BWCFiles/Ready/'
        error_directory = 'BWCFiles/Error/'
        prefix_file_path = 'BWCFiles/PropertyFile/FileNameValidationList.txt'
        validation_filename = "ExcelFileRowColumnValidator.json"
        validation_filepath = "BWCFiles/PropertyFile"
        Remarksfilepath = r"BWCFiles\PropertyFile\RemarksHeadingMapper.xlsx"
        Json_file_Codes_path = r"BWCFiles/PropertyFile\A.json"
        OutPutFilePath = r'BWCFiles/Completed/output.xlsx'
        BankDetails = r"BWCFiles\PropertyFile\BankDetails.csv"
        
        self.ask_opening_balance()
        
        self.status_var.set("Validating file names...")
        self.update_progress(10)

        fileNames = self.fileNameGetValidation(prefix_file_path, ready_directory, error_directory)

        self.status_var.set("Fetching property details...")
        self.update_progress(20)

        RemarksHeadingMapping = FileNameValidator.read_excel_file_remarks_header_map(Remarksfilepath)
        nested_list = self.read_bank_details(BankDetails)

        self.status_var.set("Processing files...")
        self.update_progress(30)

        
        for idx, fileName in enumerate(fileNames):
            self.status_var.set(f"Processing file {idx+1}/{len(fileNames)}: {fileName}")
            self.root.update()

            prop_reader = self.fileValidatorGetPropertyDetails(fileName, ready_directory, validation_filename, validation_filepath)
            if prop_reader:
                fileDetails = FileNameValidator.read_excel_file_1(os.path.join(ready_directory, fileName),prop_reader.getHeadingRowNo())
                CreditsSumsLists = FileNameValidator.GetTotalCreditAmount(fileDetails, fileName, prop_reader)
                DebitsSumLists = FileNameValidator.GetTotalDebitAmount(fileDetails, fileName, prop_reader)
                Inwardsumwithheader = FileNameValidator.RemarkTheHeadingCreditList(CreditsSumsLists, RemarksHeadingMapping, fileName)
                Outwardsumwithheader = FileNameValidator.RemarkTheHeadingDebitList(DebitsSumLists, RemarksHeadingMapping, fileName)
                Inwardcode = FileNameValidator.Code_Inwards_Generator(Inwardsumwithheader, fileName)
                Outwardcode = FileNameValidator.Code_Outwards_Generator(Outwardsumwithheader, fileName)

                
                
                FileNameValidator.dump_to_pickle(Inwardcode, Json_file_Codes_path)
                FileNameValidator.dump_to_pickle(Outwardcode, Json_file_Codes_path)

            self.update_progress(30 + 45 * (idx + 1) / len(fileNames))

        self.status_var.set("Generating consolidated Excel file...")
        self.update_progress(75)

        self.generate_consolidated_excel( nested_list, OutPutFilePath, Json_file_Codes_path)

        self.status_var.set("Consolidation complete.")
        self.update_progress(100)

        messagebox.showinfo("Info", "Consolidation complete.")

    def ask_opening_balance(self):
        self.opening_balance = float(simpledialog.askstring("Input", "Enter the opening balance:", parent=self.root))

    def update_progress(self, value):
        self.progress_var.set(value)
        self.root.update()

    def fileNameGetValidation(self, prefix_file_path, ready_directory, error_directory):
        valid_prefixes = FileNameValidator.get_valid_prefixes(prefix_file_path)
        dataframes = FileNameValidator.process_files(ready_directory, error_directory, valid_prefixes)
        filenames = [filename for filename in dataframes]
        return filenames

    def fileValidatorGetPropertyDetails(self, fileName, ready_directory, validation_filename, validation_filepath):
        input_filename = fileName
        input_filepath = ready_directory
        details = FilePraser.fetchpropertydetails(input_filename, input_filepath, validation_filename, validation_filepath)
        return details

    def read_bank_details(self, BankDetails):
        nested_list = []
        with open(BankDetails, 'r', newline='') as file:
            csv_reader = csv.reader(file)
            for row in csv_reader:
                nested_list.append(row)
        return nested_list
    

    def generate_consolidated_excel(self,nested_list, OutPutFilePath, Json_file_Codes_path):

        InwardsList = [
    "Sweep Out",
    "Dis Movement In",
    "Inter Bank Trf",
    "MF Redemption",
    "Gain on MF Redemption",
    "Equity",
    "Bank FD",
    "Interest on Bank FD",
    "WCDL/Term Loan",
    "Charge Back",
    "Settlement 1",
    "Settlement 2",
    "Settlement 3",
    "Settlement 4",
    "NPCI Revenue\n(Incl Tax)",
    "Other WLA Rev (incl Tax)",
    "MSP Rev (incl Tax)",
    "POS Rev (incl Tax)",
    "Digital Rev (incl Tax)",
    "Insurance Claim",
    "Franchisee SD",
    "ATM Cash deposit",
    "Other Misc."
]
        
        OutwardsList = [
    "Sweep Out",
    "Dis Movement Out",
    "Inter Bank Trf",
    "CW(Bank)",
    "RBI CW",
    "CW(Own CRA)",
    "CW(Retailer)",
    "CW(Alt CRA)",
    "Other",
    "Cash Dep By CRA",
    "MF investment",
    "Bank FD",
    "Term Loan Repayment",
    "Franchisee SD Refund",
    "IBM SD Paid",
    "IBM Lease Paid",
    "IBM Services Paid",
    "Other SD Paid",
    "Bank Interest",
    "POS Opex",
    "Digital Capex",
    "Digital Opex",
    "WLA Opex",
    "WLA Capex",
    "HR & ADMIN",
    "Statutory/Tax",
    "Other Misc."
]

        

        

        Headers = {
            'Details': ['Date', 'Purpose', 'Type', 'Bank', 'SAP Code', 'Opening Balance'],
            'Inwards': InwardsList,
            'Outwards':OutwardsList, 
            

                
            'Working': ['Closing Balance', 'Sanc. Limit/Min Bal', 'Effective Balance', 'Balance Tracker', 'Movement Value', 'Direct Settlement (Aq Base)', 'Total Dispense', 'Cash At ATM', 'Total Utilisation']
        }



       
        

        tuples = []
        for header, subheaders in Headers.items():
            if subheaders:
                for subheader in subheaders:
                    tuples.append((header, subheader))
            else:
                tuples.append((header, ''))

        multi_index = pd.MultiIndex.from_tuples(tuples)
        df = pd.DataFrame(columns=multi_index)
        

        Keys = FileNameValidator.load_and_delete_pickle(Json_file_Codes_path)
      
        
        s = FileNameValidator.merge_dicts(Keys)
        sorted_list = sorted(s, key=lambda x: x[('Details', 'Date')])

        
        
        Unique_Dates=set()
        Asdf=[]
        

        for i in sorted_list:
            Unique_Dates.add(i[('Details', 'Date')])
        
        row=0
        count=0
        list1=sorted(list(Unique_Dates))
       


        ClosingBalanace=[]
        CrediSum=[]
        DebitSum=[]
        sum_debits=0
        sum_credits=0

        
        OpenBalance = self.opening_balance   
        ClosingBalanace.append(OpenBalance) 
        print(sorted_list)
        
        

        for date in list1:
        
            
            
            
            
            
            for banks in nested_list:

                
                          
                 
                h={}
                h[("Details", "Date")]=date.strftime('%d %B %Y')
                h[("Details", "Purpose")] = banks[0]
                h[("Details", "Type")] = banks[1]
                h[("Details", "Bank")] = banks[2]
                h[("Details", "SAP Code")] = banks[3]       
                AAA=1
                for i in sorted_list:
                     
                     
                     if i[('Details','Date')]==date and i[('Details','Bank')]==banks[2]:
                         
                         
                         
                         sum_debits = sum(value for key, value in i.items() if key[0] == 'Outwards')
                         
                         sum_credits = sum(value for key, value in i.items() if key[0] == 'Inwards')
                         
                        
                         i[('Details','Opening Balance')]=ClosingBalanace[-1]
                         ClosingBalanaceElemenet=ClosingBalanace[-1]+sum_credits-sum_debits
                        
                         ClosingBalanace.append(ClosingBalanaceElemenet)
                         
                         i[('Working','Closing Balance')]=ClosingBalanace[-1]

                         i[("Details", "Date")]=date.strftime('%d %B %Y')
                         i[('Details','Purpose')]= banks[0]
                         i[('Details','Type')]= banks[1]
                         i[('Details','SAP Code')]= banks[3]
                         
                         df.loc[row+1]=pd.Series(i)
                         row+=1
                         AAA=0
                         
                         
                         break
                     

                if AAA==1:
                    
                    h[('Details','Opening Balance')]=ClosingBalanace[-1]
                    h[('Working','Closing Balance')]=ClosingBalanace[-1]
                    df.loc[row+1]=pd.Series(h)
                    row+=1 




            


            df.loc[row+1]=pd.Series({('Details','Date'):"Total"})

            row+=1    
            df.loc[row+1]=pd.Series({('Details','Date'):""})
            row+=1





                     
                    




                
                

                
                 
                 
                 
                     
                         
                         
           
          
                          
                       
                       
                     
                         
                    #      
                     
            #     


            # 
            # row+=1
            # df.loc[row+1]=pd.Series({"Date":""})  
            # row+=1     


                


                             
                
                        

                
            

              


                        
                        


                    

                        


                        
                    
                    
                    
                    
                    
                

            
            


        







        # ClosingBalancelist = []
        # for count, i in enumerate(sorted_list):
        #     sum_debits = sum(value for key, value in i.items() if key[0] == 'Outwards')
        #     sum_credits = sum(value for key, value in i.items() if key[0] == 'Inwards')

        #     if count == 0:
        #         OpenBalance = self.opening_balance
        #         ClosingBalance = OpenBalance + sum_credits - sum_debits
        #         ClosingBalancelist.append(ClosingBalance)
        #         for item in nested_list:
        #             if item[2] == i[('Details', 'Bank')]:
        #                 i[("Details", "Purpose")] = item[0]
        #                 i[("Details", "Type")] = item[1]
        #                 i[("Details", "SAP Code")] = item[3]
        #                 break
        #         i[("Details", "Opening Balance")] = OpenBalance
        #         i[("Working", "Closing Balance")] = ClosingBalance
        #         df.loc[count + 1] = pd.Series(i)
        #     else:
        #         OpenBalance = ClosingBalancelist[-1]
        #         ClosingBalance = OpenBalance + sum_credits - sum_debits
        #         ClosingBalancelist.clear()
        #         ClosingBalancelist.append(ClosingBalance)
        #         i[("Details", "Opening Balance")] = OpenBalance
        #         i[("Working", "Closing Balance")] = ClosingBalance
        #         for item in nested_list:
        #             if item[2] == i[('Details', 'Bank')]:
        #                 i[("Details", "Purpose")] = item[0]
        #                 i[("Details", "Type")] = item[1]
        #                 i[("Details", "SAP Code")] = item[3]
        #                 break
        #         df.loc[count + 1] = pd.Series(i)

        df.to_excel(OutPutFilePath)
        self.style_excel(OutPutFilePath)


        

    def style_excel(self, OutPutFilePath):
        wb = load_workbook(OutPutFilePath)
        ws = wb.active
        ws.delete_rows(3)

        
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[cell.column_letter].width = adjusted_width

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            max_height = 0
            for cell in row:
                try:
                    if len(str(cell.value)) > max_height:
                        max_height = len(str(cell.value))
                except:
                    pass
            ws.row_dimensions[row[0].row].height = max_height * 0.8

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

        heading_colors = {
            'Details': 'FFC000',
            'Inwards': '92D050',
            'Outwards': 'edb962',
            'Working': 'FFC000'
        }

        for row in ws.iter_rows(min_row=1, max_row=2):
            for cell in row:
                header_name = str(cell.value).strip()
                if header_name in heading_colors:
                    header_fill = PatternFill(start_color=heading_colors[header_name], end_color=heading_colors[header_name], fill_type='solid')
                    header_font = Font(color='000000', bold=True)
                    cell.fill = header_fill
                    cell.font = header_font

        if ws.max_row > 1:
            for cell in ws[2]:
                subheader_fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
                subheader_font = Font(color='000000', bold=True)
                cell.fill = subheader_fill
                cell.font = subheader_font

        wb.save(OutPutFilePath)

    def show_instructions(self):
        instructions_path = r"BWCFiles/PropertyFile/Instructions.txt"
        if os.path.exists(instructions_path):
            with open(instructions_path, 'r') as file:
                instructions = file.read()

            # Create a new window for instructions
            instruction_window = tk.Toplevel(self.root)
            instruction_window.title("Instructions")
            instruction_window.geometry("600x400")

            # Create a scrolled text widget
            scroll_text = scrolledtext.ScrolledText(instruction_window, wrap=tk.WORD)
            scroll_text.pack(expand=True, fill='both')
            scroll_text.insert(tk.INSERT, instructions)
            scroll_text.config(state=tk.DISABLED)  # Make the text read-only

            # Add a close button
            close_button = ttk.Button(instruction_window, text="Close", command=instruction_window.destroy)
            close_button.pack(pady=10)

        else:
            messagebox.showerror("Error", "Instructions file not found.")

def open_excel_file(file_path):
    try:
        os.startfile(file_path)
        print(f"Opened Excel file: {file_path}")
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
    except Exception as e:
        print(f"Error: An unexpected error occurred: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ConsolidationApp(root)
    root.mainloop()
   


    
